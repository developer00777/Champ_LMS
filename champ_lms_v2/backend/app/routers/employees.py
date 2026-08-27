"""
Admin-managed employee accounts.

Champ LMS is an internal tool: there is no public sign-up. An admin creates
each account with a team, a department and a role, and the system generates the
starting password. The employee signs in with it and is required to set their
own; the admin can read the current password at any time (see
app.core.password_vault for why that is allowed here and how it is contained).
"""
from __future__ import annotations

import csv
import io
import secrets
from datetime import datetime, timezone
from typing import Annotated

from beanie.operators import In
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from pydantic import BaseModel, EmailStr, Field

from app.core.auth import get_current_user, hash_password, require_admin, verify_password
from app.core import password_vault
from app.models.content_access import ContentAccessRule
from app.models.module import Module
from app.models.test_series import AttemptGrant, TestSeries
from app.models.user import User
from app.services import content_access
from app.services.test_attempts import attempt_status
from app.services.bunny_storage import AVATAR_BOX, bunny_storage

router = APIRouter(tags=["employees"])

# Roles an admin may assign. "ld_lead" is an admin-equivalent (see
# require_admin), so it is grantable but called out separately in the UI.
ASSIGNABLE_ROLES = ("learner", "ld_lead", "admin")

MIN_PASSWORD_LENGTH = 8

# Ambiguous characters are left out: a generated password gets read off a screen
# and typed by hand, so 0/O and 1/l/I cause avoidable support tickets.
_ALPHABET = (
    "ABCDEFGHJKLMNPQRSTUVWXYZ"
    "abcdefghijkmnopqrstuvwxyz"
    "23456789"
    "!@#$%*?"
)


def generate_password(length: int = 12) -> str:
    """
    A random starting password.

    Guarantees at least one upper, one lower, one digit and one symbol so the
    result satisfies any reasonable policy, then fills the rest randomly and
    shuffles, so the guaranteed characters aren't always in the same positions.
    """
    if length < 4:
        raise ValueError("length must be at least 4")
    pools = ("ABCDEFGHJKLMNPQRSTUVWXYZ", "abcdefghijkmnopqrstuvwxyz", "23456789", "!@#$%*?")
    chars = [secrets.choice(p) for p in pools]
    chars += [secrets.choice(_ALPHABET) for _ in range(length - len(pools))]
    # secrets-backed Fisher-Yates; random.shuffle is not cryptographically safe.
    for i in range(len(chars) - 1, 0, -1):
        j = secrets.randbelow(i + 1)
        chars[i], chars[j] = chars[j], chars[i]
    return "".join(chars)


# --------------------------------------------------------------------------
# Schemas
# --------------------------------------------------------------------------
class EmployeeCreateIn(BaseModel):
    email: EmailStr
    full_name: str = Field(min_length=1)
    employee_code: str | None = None
    department: str | None = None
    team: str | None = None
    role: str = "learner"
    # Optional: an admin may dictate the starting password instead of
    # generating one. Left unset in the normal flow.
    initial_password: str | None = None


class EmployeeUpdateIn(BaseModel):
    """Every field optional — only what's provided is changed."""
    full_name: str | None = None
    # Admin-only: employees cannot relabel themselves (see PATCH /auth/me).
    employee_code: str | None = None
    department: str | None = None
    team: str | None = None
    role: str | None = None
    is_active: bool | None = None


class ChangePasswordIn(BaseModel):
    current_password: str
    new_password: str = Field(min_length=MIN_PASSWORD_LENGTH)


def _avatar_url(user: User) -> str | None:
    """CDN URL for a profile picture, or None when none is set."""
    return bunny_storage.avatar_url(user.avatar_bunny_path)


def _normalize_code(code: str | None) -> str | None:
    """Trim and upper-case an employee code; blank becomes None."""
    code = (code or "").strip().upper()
    return code or None


async def _assert_code_free(code: str | None, *, exclude_user_id: str | None = None) -> None:
    """Employee codes identify people, so a duplicate is rejected outright."""
    if not code:
        return
    clash = await User.find_one(User.employee_code == code)
    if clash and clash.id != exclude_user_id:
        raise HTTPException(
            status_code=409,
            detail=f"Employee code {code} is already used by {clash.email}",
        )


def _employee_out(user: User, *, include_password: bool) -> dict:
    """
    Serialise an employee for the admin roster.

    `include_password` is passed explicitly by each caller rather than inferred,
    so the readable password can never leak into a non-admin response by
    accident.
    """
    out = {
        "id": user.id,
        "email": user.email,
        "full_name": user.full_name,
        "employee_code": user.employee_code,
        "avatar_url": _avatar_url(user),
        "role": user.role,
        "department": user.department,
        "team": user.team,
        "is_active": user.is_active,
        "must_change_password": user.must_change_password,
        "password_changed_at": user.password_changed_at,
        "created_at": user.created_at,
        "points": user.points,
        "level": user.level,
    }
    if include_password:
        recovered = password_vault.decrypt(user.password_recoverable)
        out["current_password"] = recovered
        # Distinguishes "we can't read it" from "there is nothing to read", so
        # the UI can tell the admin to reset rather than implying a bug.
        out["password_available"] = recovered is not None
    return out


def _apply_password(user: User, plain: str) -> None:
    """Set both the auth hash and the admin-readable copy from one plaintext."""
    user.hashed_password = hash_password(plain)
    user.password_recoverable = password_vault.encrypt(plain)


# --------------------------------------------------------------------------
# Admin: roster
# --------------------------------------------------------------------------
@router.get("/admin/employees")
async def list_employees(
    admin: Annotated[User, Depends(require_admin)],
    q: str | None = None,
    department: str | None = None,
    team: str | None = None,
    role: str | None = None,
):
    """Every account, with the current password visible to the admin."""
    users = await User.find_all().to_list()

    if q:
        needle = q.strip().lower()
        users = [
            u for u in users
            if needle in (u.email or "").lower()
            or needle in (u.full_name or "").lower()
        ]
    if department:
        users = [u for u in users if (u.department or "") == department]
    if team:
        users = [u for u in users if (u.team or "") == team]
    if role:
        users = [u for u in users if u.role == role]

    users.sort(key=lambda u: (u.full_name or u.email or "").lower())
    return {
        "employees": [_employee_out(u, include_password=True) for u in users],
        "total": len(users),
        "departments": sorted({u.department for u in users if u.department}),
        "teams": sorted({u.team for u in users if u.team}),
    }


@router.post("/admin/employees", status_code=201)
async def create_employee(
    body: EmployeeCreateIn,
    admin: Annotated[User, Depends(require_admin)],
):
    """
    Provision an employee account. Returns the starting password so the admin
    can pass it on; it stays readable on the roster afterwards.
    """
    if body.role not in ASSIGNABLE_ROLES:
        raise HTTPException(
            status_code=422,
            detail=f"role must be one of: {', '.join(ASSIGNABLE_ROLES)}",
        )

    email = body.email.strip().lower()
    if await User.find_one(User.email == email):
        raise HTTPException(status_code=409, detail="That email already has an account")

    code = _normalize_code(body.employee_code)
    await _assert_code_free(code)

    if body.initial_password is not None:
        if len(body.initial_password) < MIN_PASSWORD_LENGTH:
            raise HTTPException(
                status_code=422,
                detail=f"Password must be at least {MIN_PASSWORD_LENGTH} characters",
            )
        starting = body.initial_password
    else:
        starting = generate_password()

    user = User(
        email=email,
        full_name=body.full_name.strip(),
        hashed_password="",  # replaced by _apply_password below
        role=body.role,
        department=(body.department or "").strip() or None,
        team=(body.team or "").strip() or None,
        employee_code=code,
        must_change_password=True,
        created_by_admin_id=admin.id,
    )
    _apply_password(user, starting)
    await user.insert()

    return {
        **_employee_out(user, include_password=True),
        "initial_password": starting,
    }


def _read_tabular(data: bytes, filename: str) -> list[dict]:
    """
    Read a CSV or Excel master tracker into row dicts keyed by our field names.

    Excel is handled with openpyxl (already a dependency of nothing else here,
    so it degrades to a clear message if absent). CSV is decoded tolerantly:
    trackers exported from Excel are frequently cp1252 or UTF-8-BOM rather than
    clean UTF-8, and failing on an encoding quirk would be a poor experience.
    """
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        return _read_excel(data)
    if name.endswith(".xls"):
        raise HTTPException(
            status_code=415,
            detail="Old .xls files aren't supported. Save as .xlsx or CSV.",
        )
    return _read_csv(data)


def _read_csv(data: bytes) -> list[dict]:
    text = None
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise HTTPException(status_code=422, detail="Could not decode this file as text")

    try:
        # Sniff the delimiter: trackers are exported as comma- or
        # semicolon-separated depending on the machine's locale.
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    mapping = _map_headers(reader.fieldnames or [])
    if not mapping:
        raise HTTPException(
            status_code=422,
            detail=(
                "Couldn't find the expected columns. Include headings for "
                "employee ID, employee name and official email ID."
            ),
        )
    return [
        {mapping[k]: (v or "") for k, v in row.items() if k in mapping}
        for row in reader
    ]


def _read_excel(data: bytes) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is pinned
        raise HTTPException(
            status_code=415,
            detail="Excel support unavailable: the 'openpyxl' package is not installed.",
        ) from exc

    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Could not read this workbook: {exc}") from exc

    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return []

    headers = [str(h).strip() if h is not None else "" for h in header]
    mapping = _map_headers(headers)
    if not mapping:
        raise HTTPException(
            status_code=422,
            detail=(
                "Couldn't find the expected columns. Include headings for "
                "employee ID, employee name and official email ID."
            ),
        )

    out: list[dict] = []
    for raw_row in rows:
        row: dict[str, str] = {}
        for header_name, value in zip(headers, raw_row):
            field = mapping.get(header_name)
            if not field:
                continue
            # Excel stores a numeric employee code as a float (1042.0), which
            # would otherwise be imported with a stray decimal.
            if isinstance(value, float) and value.is_integer():
                value = int(value)
            row[field] = "" if value is None else str(value).strip()
        if any(row.values()):
            out.append(row)
    return out


# Header aliases accepted in a bulk upload, so a real master tracker doesn't
# have to be renamed to match us exactly.
_BULK_HEADERS = {
    "employee_code": {
        "employee id", "employee code", "emp id", "emp code", "employee_id",
        "employee_code", "empid", "code", "id",
    },
    "full_name": {
        "employee name", "full name", "name", "employee_name", "full_name",
    },
    "email": {
        "official email id", "official email", "email id", "email", "e-mail",
        "official_email", "email_address", "mail",
    },
    "department": {"department", "dept"},
    "team": {"team", "squad"},
    "role": {"role", "privilege", "privileges", "access"},
}

MAX_BULK_ROWS = 2000


def _map_headers(fieldnames: list[str]) -> dict[str, str]:
    """Map a tracker's column names onto our fields, case-insensitively."""
    mapping: dict[str, str] = {}
    for raw in fieldnames or []:
        key = (raw or "").strip().lower().replace("_", " ")
        for field, aliases in _BULK_HEADERS.items():
            if key in aliases or key.replace(" ", "_") in aliases:
                mapping[raw] = field
                break
    return mapping


@router.post("/admin/employees/bulk-upload")
async def bulk_upload_employees(
    admin: Annotated[User, Depends(require_admin)],
    file: UploadFile = File(...),
    dry_run: bool = False,
):
    """
    Create many accounts at once from a master tracker (CSV or Excel).

    Expects employee ID, employee name and official email columns; department,
    team and role are optional. Column names are matched loosely, so the usual
    tracker headings work without editing the file.

    Every row is validated before anything is written, and the whole upload is
    rejected if any row is bad. A tracker half-imported is worse than one not
    imported: the admin cannot tell which people already exist, and re-running
    it produces a pile of duplicate-email errors.

    `dry_run` validates and reports without creating anything.
    """
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=422, detail="The uploaded file is empty")
    if len(raw) > 5 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File too large. Limit is 5MB.")

    rows = _read_tabular(raw, file.filename or "")
    if not rows:
        raise HTTPException(
            status_code=422,
            detail="No rows found. Include a header row and at least one employee.",
        )
    if len(rows) > MAX_BULK_ROWS:
        raise HTTPException(
            status_code=413,
            detail=f"{len(rows)} rows exceeds the {MAX_BULK_ROWS}-row limit.",
        )

    existing_emails = {u.email for u in await User.find_all().to_list()}
    existing_codes = {u.employee_code for u in await User.find_all().to_list() if u.employee_code}

    valid: list[dict] = []
    errors: list[dict] = []
    seen_emails: set[str] = set()
    seen_codes: set[str] = set()

    for i, row in enumerate(rows, start=2):  # row 1 is the header, so data starts at 2
        email = (row.get("email") or "").strip().lower()
        name = (row.get("full_name") or "").strip()
        code = _normalize_code(row.get("employee_code"))
        role = (row.get("role") or "learner").strip().lower() or "learner"

        def bad(msg: str) -> None:
            errors.append({"row": i, "email": email or None, "error": msg})

        if not email and not name and not code:
            continue  # blank spacer row in the spreadsheet
        if not email:
            bad("Missing official email ID")
            continue
        if "@" not in email or email.startswith("@") or email.endswith("@"):
            bad(f"'{email}' is not a valid email address")
            continue
        if not name:
            bad("Missing employee name")
            continue
        if role not in ASSIGNABLE_ROLES:
            bad(f"Unknown role '{role}'. Valid: {', '.join(ASSIGNABLE_ROLES)}")
            continue
        if email in existing_emails:
            bad("An account with this email already exists")
            continue
        if email in seen_emails:
            bad("Duplicate email within the file")
            continue
        if code and code in existing_codes:
            bad(f"Employee code {code} is already used")
            continue
        if code and code in seen_codes:
            bad(f"Duplicate employee code {code} within the file")
            continue

        seen_emails.add(email)
        if code:
            seen_codes.add(code)
        valid.append({
            "email": email, "full_name": name, "employee_code": code,
            "department": (row.get("department") or "").strip() or None,
            "team": (row.get("team") or "").strip() or None,
            "role": role,
        })

    if errors:
        raise HTTPException(
            status_code=422,
            detail={
                "message": (
                    f"{len(errors)} row(s) could not be imported, so nothing was "
                    "created. Fix them and upload again."
                ),
                "errors": errors[:50],
                "error_count": len(errors),
                "valid_count": len(valid),
            },
        )

    if dry_run:
        return {
            "dry_run": True,
            "would_create": len(valid),
            "employees": [
                {k: v for k, v in row.items() if k != "role"} | {"role": row["role"]}
                for row in valid
            ],
        }

    created = []
    for row in valid:
        password = generate_password()
        user = User(
            email=row["email"],
            full_name=row["full_name"],
            employee_code=row["employee_code"],
            hashed_password="",  # set by _apply_password
            role=row["role"],
            department=row["department"],
            team=row["team"],
            must_change_password=True,
            created_by_admin_id=admin.id,
        )
        _apply_password(user, password)
        await user.insert()
        created.append({
            "id": user.id,
            "employee_code": user.employee_code,
            "full_name": user.full_name,
            "email": user.email,
            "department": user.department,
            "team": user.team,
            "role": user.role,
            "initial_password": password,
        })

    return {"dry_run": False, "created_count": len(created), "created": created}


@router.patch("/admin/employees/{user_id}")
async def update_employee(
    user_id: str,
    body: EmployeeUpdateIn,
    admin: Annotated[User, Depends(require_admin)],
):
    """Change an employee's team, department, role or active state."""
    user = await User.get(user_id)
    if not user:
        raise HTTPException(status_code=404, detail="Employee not found")

    if body.role is not None and body.role not in ASSIGNABLE_ROLES:
        raise HTTPException(
            status_code=422,
            detail=f"role must be one of: {', '.join(ASSIGNABLE_ROLES)}",
        )

    # Guard against an admin removing their own access and locking the tool.
    if user.id == admin.id:
        if body.role is not None and body.role not in ("admin", "ld_lead"):
            raise HTTPException(
                status_code=422,
                detail="You cannot remove your own admin access.",
            )
        if body.is_active is False:
            raise HTTPException(
                status_code=422, detail="You cannot deactivate your own account."
            )

    if body.employee_code is not None:
        code = _normalize_code(body.employee_code)
        await _assert_code_free(code, exclude_user_id=user.id)
        user.employee_code = code
    if body.full_name is not None:
        user.full_name = body.full_name.strip() or None
    if body.department is not None:
        user.department = body.department.strip() or None
    if body.team is not None:
        user.team = body.team.strip() or None
    if body.role is not None:
        user.role = body.role
    if body.is_active is not None:
        user.is_active = body.is_active

    await user.save()
    return _employee_out(user, include_password=True)


@router.post("/admin/employees/{user_id}/reset-password")
async def reset_employee_password(
    user_id: str,
    admin: Annotated[User, Depends(require_admin)],
):
    """
    Issue a fresh random password and require the employee to change it again.
    Used when someone is locked out or a password should be invalidated.
    """
    user = await User.get(user_id)
    if not user:
        raise HTTPException(status_code=404, detail="Employee not found")

    new_password = generate_password()
    _apply_password(user, new_password)
    user.must_change_password = True
    user.password_changed_at = None
    await user.save()

    return {
        **_employee_out(user, include_password=True),
        "initial_password": new_password,
    }


@router.delete("/admin/employees/{user_id}")
async def delete_employee(
    user_id: str,
    admin: Annotated[User, Depends(require_admin)],
):
    """
    Deactivate an account. Deliberately not a hard delete: attempts, progress
    and leaderboard history reference the user, so removing the row would leave
    those dangling.
    """
    user = await User.get(user_id)
    if not user:
        raise HTTPException(status_code=404, detail="Employee not found")
    if user.id == admin.id:
        raise HTTPException(
            status_code=422, detail="You cannot deactivate your own account."
        )

    user.is_active = False
    await user.save()
    return {"id": user.id, "is_active": user.is_active}


# --------------------------------------------------------------------------
# Employee: own profile
# --------------------------------------------------------------------------
MAX_AVATAR_BYTES = 5 * 1024 * 1024
_AVATAR_TYPES = {
    "image/jpeg": ".jpg", "image/jpg": ".jpg", "image/png": ".png",
    "image/webp": ".webp", "image/gif": ".gif",
}


class ProfileUpdateIn(BaseModel):
    """
    What an employee may change about themselves.

    Deliberately excludes employee_code, role, team and department: those are
    org facts an admin owns, and letting people edit their own would make the
    roster untrustworthy.
    """
    full_name: str | None = None


@router.patch("/auth/me")
async def update_own_profile(
    body: ProfileUpdateIn,
    user: Annotated[User, Depends(get_current_user)],
):
    """Update the caller's own editable profile fields."""
    if body.full_name is not None:
        name = body.full_name.strip()
        if not name:
            raise HTTPException(status_code=422, detail="Name cannot be empty")
        user.full_name = name
        await user.save()
    return {
        "id": user.id,
        "full_name": user.full_name,
        "employee_code": user.employee_code,
        "avatar_url": _avatar_url(user),
    }


@router.post("/auth/me/avatar")
async def upload_own_avatar(
    user: Annotated[User, Depends(get_current_user)],
    file: UploadFile = File(...),
):
    """
    Upload the caller's profile picture to Bunny Storage.

    The stored path includes the user id so one person's picture can never
    overwrite another's, and a timestamp so a replacement gets a fresh URL
    rather than being masked by a stale CDN cache entry.
    """
    content_type = (file.content_type or "").lower().split(";")[0].strip()
    ext = _AVATAR_TYPES.get(content_type)
    if not ext:
        raise HTTPException(
            status_code=415,
            detail="Upload a JPEG, PNG, WebP or GIF image.",
        )

    data = await file.read()
    if not data:
        raise HTTPException(status_code=422, detail="The uploaded image is empty")
    if len(data) > MAX_AVATAR_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"Image too large ({len(data) // 1024 // 1024}MB). Limit is 5MB.",
        )

    stamp = int(datetime.now(timezone.utc).timestamp())
    previous = user.avatar_bunny_path

    # Resized and re-encoded here, once, rather than per request via Bunny
    # Optimizer — see the cost note in services/bunny_storage.py. The final
    # extension comes from the encode, so the path is built without `ext`.
    try:
        path = await bunny_storage.upload_optimized(
            f"avatars/{user.id}/{stamp}", data, AVATAR_BOX
        )
    except Exception as exc:  # noqa: BLE001 - surface a usable message
        raise HTTPException(
            status_code=502, detail=f"Could not upload the image: {exc}"
        ) from exc

    user.avatar_bunny_path = path
    await user.save()

    # Best-effort cleanup of the replaced file. A failure here must not fail the
    # request: the new picture is already live and correct.
    if previous and previous != path:
        try:
            await bunny_storage.delete_thumbnail(previous)
        except Exception:  # noqa: BLE001
            pass

    return {"avatar_url": _avatar_url(user), "avatar_bunny_path": path}


@router.delete("/auth/me/avatar")
async def delete_own_avatar(user: Annotated[User, Depends(get_current_user)]):
    """Remove the caller's profile picture, falling back to initials."""
    path = user.avatar_bunny_path
    user.avatar_bunny_path = None
    await user.save()
    if path:
        try:
            await bunny_storage.delete_thumbnail(path)
        except Exception:  # noqa: BLE001 - the row is already cleared
            pass
    return {"avatar_url": None}


# --------------------------------------------------------------------------
# Employee: change own password
# --------------------------------------------------------------------------
@router.post("/auth/change-password")
async def change_own_password(
    body: ChangePasswordIn,
    user: Annotated[User, Depends(get_current_user)],
):
    """
    Replace the caller's own password.

    Requires the current password, so a stolen token alone can't lock the
    real employee out of their account.
    """
    if not verify_password(body.current_password, user.hashed_password):
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    if body.new_password == body.current_password:
        raise HTTPException(
            status_code=422, detail="New password must be different from the current one"
        )

    _apply_password(user, body.new_password)
    user.must_change_password = False
    user.password_changed_at = datetime.now(timezone.utc)
    await user.save()

    return {
        "id": user.id,
        "must_change_password": user.must_change_password,
        "password_changed_at": user.password_changed_at,
    }


# --------------------------------------------------------------------------
# Admin: content targeting
#
# Two levers, described in app/services/content_access.py:
#   * audience         — teams / departments / roles that can see a piece of
#                        content, and the teams it is required of.
#   * per-person rule  — grant, require or revoke for one individual.
#
# Both levers apply to two kinds of content — learning modules and test series.
# The endpoints below are written once over a `kind` and mounted under both
# /content-access/modules/... and /content-access/tests/..., because an admin
# uses the same screen for both and two copies of this logic is how the two
# drift apart.
# --------------------------------------------------------------------------
class ModuleAudienceIn(BaseModel):
    """
    Set content's audience. Passing [] clears a dimension (opening it up);
    omitting a field leaves that dimension untouched.
    """
    audience_teams: list[str] | None = None
    audience_departments: list[str] | None = None
    target_roles: list[str] | None = None
    required_for_teams: list[str] | None = None


class AccessRuleIn(BaseModel):
    user_id: str
    access: str  # grant | required | revoke
    reason: str | None = None


class GrantAttemptsIn(BaseModel):
    """Give one person extra attempts on one test."""
    user_id: str
    # How many extra attempts to add. Capped low deliberately: handing someone
    # 500 retakes is almost always a typo, and an admin who genuinely wants the
    # test uncapped for everyone should clear max_attempts instead.
    extra_attempts: int = Field(default=1, ge=1, le=20)
    reason: str | None = None


def _clean_list(values: list[str] | None) -> list[str] | None:
    """Trim, drop blanks, de-duplicate case-insensitively, keep order."""
    if values is None:
        return None
    seen: set[str] = set()
    out: list[str] = []
    for v in values:
        v = (v or "").strip()
        if v and v.lower() not in seen:
            seen.add(v.lower())
            out.append(v)
    return out


# The two kinds of content this section administers, and how to load each.
_KIND_MODULE = ContentAccessRule.KIND_MODULE
_KIND_TEST = ContentAccessRule.KIND_TEST
_CONTENT_DOCS = {_KIND_MODULE: Module, _KIND_TEST: TestSeries}
_NOT_FOUND = {_KIND_MODULE: "Module not found", _KIND_TEST: "Test series not found"}


async def _get_content_or_404(kind: str, content_id: str):
    doc = await _CONTENT_DOCS[kind].get(content_id)
    if not doc:
        raise HTTPException(status_code=404, detail=_NOT_FOUND[kind])
    return doc


def _audience_out(content, kind: str = _KIND_MODULE) -> dict:
    """
    One module or test with its audience.

    `id` is kept as the key name for both so the admin screen renders either
    without a branch; `kind` tells it which endpoints to call back into.
    """
    out = {
        "id": content.id,
        "kind": kind,
        "title": content.title,
        "category": content.category,
        "is_published": content.is_published,
        "audience_teams": content.audience_teams or [],
        "audience_departments": content.audience_departments or [],
        "target_roles": content.target_roles or [],
        "required_for_teams": content.required_for_teams or [],
        "is_restricted": content_access.is_restricted(content),
    }
    if kind == _KIND_TEST:
        # Extra context the admin needs to judge a test row: the legacy
        # single-department targeting (which still counts as an audience), and
        # whether the paper is actually sittable.
        out["department"] = content.department
        out["total_questions"] = len(content.questions)
    return out


async def _content_people(kind: str, content_id: str) -> dict:
    """
    Who can currently open this module/test, and why.

    For a test this also carries each person's attempt position, so the one
    screen that decides who may sit a test is also where an admin hands back a
    retake — rather than making them find the same person again elsewhere.
    """
    content = await _get_content_or_404(kind, content_id)

    users = await User.find_all().to_list()
    rules = await ContentAccessRule.find(
        ContentAccessRule.module_id == content_id,
        ContentAccessRule.content_kind == kind,
    ).to_list()
    by_user = {r.user_id: r for r in rules}

    people = []
    for u in sorted(users, key=lambda x: (x.full_name or x.email or "").lower()):
        rule = by_user.get(u.id)
        override = {content_id: rule.access} if rule else {}
        allowed = content_access.can_access(content, u, override)
        if content_access.is_admin(u):
            why = "admin"
        elif rule:
            why = f"rule: {rule.access}"
        elif content_access.matches_audience(content, u):
            why = "audience" if content_access.is_restricted(content) else "open to everyone"
        else:
            why = "not in audience"
        row = {
            "user_id": u.id,
            "full_name": u.full_name,
            "email": u.email,
            "team": u.team,
            "department": u.department,
            "role": u.role,
            "is_active": u.is_active,
            "can_access": allowed,
            "required": content_access.is_required(content, u, override),
            "rule": rule.access if rule else None,
            "reason": rule.reason if rule else None,
            "why": why,
        }
        if kind == _KIND_TEST:
            # used / granted_extra / allowed / left / exhausted for this person.
            row["attempts"] = await attempt_status(content, u.id)
        people.append(row)

    out = {
        **_audience_out(content, kind),
        "people": people,
        "can_access_count": sum(1 for p in people if p["can_access"]),
    }
    if kind == _KIND_TEST:
        # None = uncapped, in which case granting extra attempts is meaningless
        # and the UI hides the control rather than offering a no-op.
        out["max_attempts"] = content.max_attempts
    return out


async def _set_audience(kind: str, content_id: str, body: ModuleAudienceIn) -> dict:
    """Restrict a module/test to teams/departments/roles, or open it up again."""
    content = await _get_content_or_404(kind, content_id)

    if body.target_roles is not None:
        bad = [r for r in body.target_roles if r not in ASSIGNABLE_ROLES]
        if bad:
            raise HTTPException(
                status_code=422,
                detail=f"Unknown role(s): {', '.join(bad)}. Valid: {', '.join(ASSIGNABLE_ROLES)}",
            )

    if body.audience_teams is not None:
        content.audience_teams = _clean_list(body.audience_teams)
    if body.audience_departments is not None:
        content.audience_departments = _clean_list(body.audience_departments)
    if body.target_roles is not None:
        content.target_roles = _clean_list(body.target_roles)
    if body.required_for_teams is not None:
        content.required_for_teams = _clean_list(body.required_for_teams)

    await content.save()
    return _audience_out(content, kind)


async def _set_person_rule(
    kind: str, content_id: str, body: AccessRuleIn, admin: User
) -> dict:
    """Grant, require or revoke one piece of content for one person."""
    valid = (ContentAccessRule.GRANT, ContentAccessRule.REQUIRED, ContentAccessRule.REVOKE)
    if body.access not in valid:
        raise HTTPException(
            status_code=422, detail=f"access must be one of: {', '.join(valid)}"
        )

    await _get_content_or_404(kind, content_id)
    target = await User.get(body.user_id)
    if not target:
        raise HTTPException(status_code=404, detail="Employee not found")

    existing = await ContentAccessRule.find_one(
        ContentAccessRule.user_id == body.user_id,
        ContentAccessRule.module_id == content_id,
    )
    if existing:
        existing.access = body.access
        existing.reason = body.reason
        existing.created_by_admin_id = admin.id
        # Repair the kind on a rule written before test series were covered.
        existing.content_kind = kind
        await existing.save()
        rule = existing
    else:
        rule = ContentAccessRule(
            user_id=body.user_id,
            module_id=content_id,
            content_kind=kind,
            access=body.access,
            reason=body.reason,
            created_by_admin_id=admin.id,
        )
        await rule.insert()

    return {
        "content_id": content_id,
        "kind": kind,
        # Kept for the module screen, which has always read module_id.
        "module_id": content_id,
        "user_id": rule.user_id,
        "access": rule.access,
        "reason": rule.reason,
    }


async def _clear_person_rule(kind: str, content_id: str, user_id: str) -> dict:
    """Drop a per-person rule, so the audience rules apply again."""
    rule = await ContentAccessRule.find_one(
        ContentAccessRule.user_id == user_id,
        ContentAccessRule.module_id == content_id,
    )
    if not rule:
        raise HTTPException(status_code=404, detail="No rule set for that person")
    await rule.delete()
    return {
        "content_id": content_id,
        "kind": kind,
        "module_id": content_id,
        "user_id": user_id,
        "access": None,
    }


@router.get("/admin/content-access/modules")
async def list_module_audiences(admin: Annotated[User, Depends(require_admin)]):
    """
    Every module and every test series with its audience, plus the
    teams/departments available to assign (taken from real employee records, so
    the admin picks from what actually exists rather than retyping names).

    Tests are returned alongside modules because they are content an admin
    targets the same way. `tests` is a separate key rather than being mixed into
    `modules` so the screen can show them as their own section — and so an older
    client that only reads `modules` keeps working unchanged.
    """
    modules = await Module.find_all().sort(-Module.created_at).to_list()
    tests = await TestSeries.find_all().sort(-TestSeries.created_at).to_list()
    users = await User.find_all().to_list()
    return {
        "modules": [_audience_out(m, _KIND_MODULE) for m in modules],
        "tests": [_audience_out(t, _KIND_TEST) for t in tests],
        "teams": sorted({u.team for u in users if u.team}),
        "departments": sorted({u.department for u in users if u.department}),
        "roles": list(ASSIGNABLE_ROLES),
    }


@router.patch("/admin/content-access/modules/{module_id}")
async def set_module_audience(
    module_id: str,
    body: ModuleAudienceIn,
    admin: Annotated[User, Depends(require_admin)],
):
    """Restrict a module to teams/departments/roles, or open it up again."""
    return await _set_audience(_KIND_MODULE, module_id, body)


@router.get("/admin/content-access/modules/{module_id}/people")
async def module_access_people(
    module_id: str,
    admin: Annotated[User, Depends(require_admin)],
):
    """
    Who can currently see this module and why — audience match, an explicit
    per-person rule, or admin bypass. Lets an admin verify the effect of a
    restriction instead of guessing.
    """
    return await _content_people(_KIND_MODULE, module_id)


@router.put("/admin/content-access/modules/{module_id}/people")
async def set_person_access(
    module_id: str,
    body: AccessRuleIn,
    admin: Annotated[User, Depends(require_admin)],
):
    """Grant, require or revoke this module for one person."""
    return await _set_person_rule(_KIND_MODULE, module_id, body, admin)


@router.delete("/admin/content-access/modules/{module_id}/people/{user_id}")
async def clear_person_access(
    module_id: str,
    user_id: str,
    admin: Annotated[User, Depends(require_admin)],
):
    """Drop a per-person rule, so the module's audience rules apply again."""
    return await _clear_person_rule(_KIND_MODULE, module_id, user_id)


# --- the same four levers, for test series -------------------------------
# Separate paths rather than a `kind` query parameter: the ids live in
# different collections, so a typo'd id should 404 as "test not found" rather
# than silently resolving to a module with the same id shape.
@router.patch("/admin/content-access/tests/{test_id}")
async def set_test_audience(
    test_id: str,
    body: ModuleAudienceIn,
    admin: Annotated[User, Depends(require_admin)],
):
    """Restrict a test series to teams/departments/roles, or open it up again."""
    return await _set_audience(_KIND_TEST, test_id, body)


@router.get("/admin/content-access/tests/{test_id}/people")
async def test_access_people(
    test_id: str,
    admin: Annotated[User, Depends(require_admin)],
):
    """Who can currently sit this test, and why."""
    return await _content_people(_KIND_TEST, test_id)


@router.put("/admin/content-access/tests/{test_id}/people")
async def set_test_person_access(
    test_id: str,
    body: AccessRuleIn,
    admin: Annotated[User, Depends(require_admin)],
):
    """Grant, require or revoke this test for one person."""
    return await _set_person_rule(_KIND_TEST, test_id, body, admin)


@router.delete("/admin/content-access/tests/{test_id}/people/{user_id}")
async def clear_test_person_access(
    test_id: str,
    user_id: str,
    admin: Annotated[User, Depends(require_admin)],
):
    """Drop a per-person rule, so the test's audience rules apply again."""
    return await _clear_person_rule(_KIND_TEST, test_id, user_id)


# --- extra attempts (retakes) -------------------------------------------
# Mounted here rather than under /admin/test-series because deciding who may
# sit a test and handing one person another go at it are the same decision,
# made about the same person, on the same screen.
@router.post("/admin/content-access/tests/{test_id}/grants", status_code=201)
async def grant_extra_attempts(
    test_id: str,
    body: GrantAttemptsIn,
    admin: Annotated[User, Depends(require_admin)],
):
    """
    Give one person extra attempts on one test.

    Additive: granting twice gives two extra attempts, which is what an admin
    means when they click the button a second time after the learner burned the
    first grant. Refused on an uncapped test — there is nothing to extend, and
    storing the row would imply a limit that does not exist.
    """
    test = await _get_content_or_404(_KIND_TEST, test_id)
    if test.max_attempts is None:
        raise HTTPException(
            status_code=422,
            detail="This test has unlimited attempts — there is nothing to grant.",
        )

    target = await User.get(body.user_id)
    if not target:
        raise HTTPException(status_code=404, detail="Employee not found")

    grant = AttemptGrant(
        test_id=test.id,
        user_id=target.id,
        extra_attempts=body.extra_attempts,
        reason=(body.reason or "").strip() or None,
        granted_by=admin.id,
    )
    await grant.insert()

    return {
        "grant_id": grant.id,
        "test_id": test.id,
        "user_id": target.id,
        "full_name": target.full_name,
        "extra_attempts": grant.extra_attempts,
        "reason": grant.reason,
        "granted_at": grant.granted_at,
        "attempts": await attempt_status(test, target.id),
    }


@router.get("/admin/content-access/tests/{test_id}/grants")
async def list_extra_attempt_grants(
    test_id: str, admin: Annotated[User, Depends(require_admin)]
):
    """Every grant issued on this test, newest first, with names resolved."""
    test = await _get_content_or_404(_KIND_TEST, test_id)
    grants = await AttemptGrant.find(AttemptGrant.test_id == test.id).sort(
        -AttemptGrant.granted_at
    ).to_list()

    ids = list({g.user_id for g in grants} | {g.granted_by for g in grants})
    users = {u.id: u for u in await User.find(In(User.id, ids)).to_list()} if ids else {}

    return [
        {
            "grant_id": g.id,
            "user_id": g.user_id,
            "full_name": (users[g.user_id].full_name if g.user_id in users else None),
            "extra_attempts": g.extra_attempts,
            "reason": g.reason,
            "granted_by": g.granted_by,
            "granted_by_name": (
                users[g.granted_by].full_name if g.granted_by in users else None
            ),
            "granted_at": g.granted_at,
        }
        for g in grants
    ]


@router.delete("/admin/content-access/tests/{test_id}/grants/{grant_id}")
async def revoke_extra_attempt_grant(
    test_id: str,
    grant_id: str,
    admin: Annotated[User, Depends(require_admin)],
):
    """
    Withdraw a grant issued by mistake.

    Only removes the unused allowance — attempts the learner already sat stay
    on their record. Deleting a grant they have spent lowers their ceiling
    below what they used, which the take/submit checks read simply as "out of
    attempts"; it never invalidates a score.
    """
    grant = await AttemptGrant.get(grant_id)
    if not grant or grant.test_id != test_id:
        raise HTTPException(status_code=404, detail="Grant not found on this test")
    user_id = grant.user_id
    await grant.delete()

    test = await _get_content_or_404(_KIND_TEST, test_id)
    return {"revoked": grant_id, "attempts": await attempt_status(test, user_id)}


@router.get("/admin/content-access/employees/{user_id}")
async def employee_access_overview(
    user_id: str,
    admin: Annotated[User, Depends(require_admin)],
):
    """
    The catalogue from one employee's point of view — what they can open, what
    is required of them, and where any explicit rule applies.
    """
    target = await User.get(user_id)
    if not target:
        raise HTTPException(status_code=404, detail="Employee not found")

    modules = await Module.find_all().sort(-Module.created_at).to_list()
    tests = await TestSeries.find_all().sort(-TestSeries.created_at).to_list()
    # Rules are read per kind so a module rule is never applied to a test.
    overrides = await content_access.rules_for_user(user_id)
    test_overrides = await content_access.test_rules_for_user(user_id)

    items = [
        {
            "module_id": m.id,
            "title": m.title,
            "category": m.category,
            "is_published": m.is_published,
            "is_restricted": content_access.is_restricted(m),
            "can_access": content_access.can_access(m, target, overrides),
            "required": content_access.is_required(m, target, overrides),
            "rule": overrides.get(m.id),
        }
        for m in modules
    ]
    test_items = [
        {
            "test_id": t.id,
            "title": t.title,
            "category": t.category,
            "is_published": t.is_published,
            "is_restricted": content_access.is_restricted(t),
            "can_access": content_access.can_access(t, target, test_overrides),
            "required": content_access.is_required(t, target, test_overrides),
            "rule": test_overrides.get(t.id),
        }
        for t in tests
    ]
    return {
        "user_id": target.id,
        "full_name": target.full_name,
        "email": target.email,
        "team": target.team,
        "department": target.department,
        "role": target.role,
        "modules": items,
        "tests": test_items,
        # Counts stay module-only so an existing caller's numbers don't change
        # meaning; the test totals are their own keys.
        "accessible_count": sum(1 for i in items if i["can_access"]),
        "required_count": sum(1 for i in items if i["required"]),
        "tests_accessible_count": sum(1 for i in test_items if i["can_access"]),
        "tests_required_count": sum(1 for i in test_items if i["required"]),
    }
